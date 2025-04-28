# ---
# jupyter:
#   jupytext:
#     formats: ipynb,py:percent
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#       jupytext_version: 1.17.0
#   kernelspec:
#     display_name: Python 3 (ipykernel)
#     language: python
#     name: python3
# ---

# %%
# Import configuration
import json
from datetime import datetime

def load_config(config_path="config.json"):
    """Load configuration from a JSON file."""
    with open(config_path, 'r') as file:
        config = json.load(file)
    return config["sqlserver_name"], config["sqlserver_db"], config["sqlserver_ip"], config["sqlserver_port"], config["sqlserver_user"], config["sqlserver_pwd"], config['base_url'], config['username'], config['password'], config['output_directory'], config['cpd_directory']
    

# Test loading configuration
sqlserver_name, sqlserver_db, sqlserver_ip, sqlserver_port, sqlserver_user, sqlserver_pwd, base_url, username, password, output_dir, cpd_directory = load_config()
print("Configuration loaded successfully.")

# %%
# Establish a database server connection
import pyodbc
from sqlalchemy import create_engine

conn = """
    Driver={{ODBC Driver 17 for SQL Server}};
    Server={},{};
    Database={};
    authentication=SqlPassword;UID={};PWD={};
    TrustServerCertificate=yes;
    """.format(sqlserver_ip, sqlserver_port, sqlserver_db, sqlserver_user, sqlserver_pwd)

sql_conn = pyodbc.connect(conn, autocommit=False)

cursor = sql_conn.cursor()
cursor.execute('SELECT schNo, schName FROM Schools')

for i, row in enumerate(cursor):
    if i >= 3:
        break
    print(row)

# %%
# Find excel workbook to load
import os

# Find CPD sample files to upload, skipping unwanted ones
def find_sample_files(directory, skip_prefix="CPD-source-data-workbook", extension=".xlsx"):
    """Scan directory and list files NOT starting with skip_prefix, matching extension."""
    files = []
    for filename in os.listdir(directory):
        if filename.endswith(extension) and not filename.startswith(skip_prefix):
            files.append(os.path.join(directory, filename))
    return files

# Load sample files
sample_files = find_sample_files(cpd_directory)

print(f"Found {len(sample_files)} sample files to upload:")
for f in sample_files:
    print("-", os.path.basename(f))

# %%
# Load all excel workbook in memory.
import openpyxl

# Load all Excel files into memory (workbooks)
def load_excel_workbooks(file_list):
    """Load Excel workbooks into memory."""
    workbooks = {}
    for file_path in file_list:
        try:
            wb = openpyxl.load_workbook(file_path)
            workbooks[file_path] = wb
            print(f"✅ Loaded: {os.path.basename(file_path)}")
        except Exception as e:
            print(f"❌ Failed to load {os.path.basename(file_path)}: {e}")
    return workbooks

# Actually load them now
all_workbooks = load_excel_workbooks(sample_files)

print(f"\nTotal workbooks loaded: {len(all_workbooks)}")

# %%
# Convert Excel workbook data into XML
import xml.etree.ElementTree as ET

def workbook_to_xml(wb):
    ws = wb["CPD data"]  # your sheet is called exactly "CPD data"

    # Build a list of column headers
    headers = []
    for cell in ws[1]:
        if cell.value is not None:
            headers.append(cell.value.strip())
        else:
            headers.append("")

    # Column mapping (your provided one)
    column_mapping = {
        "CPD Name": "CPDName",
        "CPD Format": "CPDFormat",
        "CPD Focus": "CPDFocus",
        "Location": "Location",
        "Year": "Year",
        "Start Date (YYY-MM-DD)": "StartDate",
        "End Date (YYY-MM-DD)": "EndDate",
        "Duration in Days": "DurationDays",
        "Duration in Hours": "DurationHours",
        "Teacher PF Number": "TeacherPFNumber",
        "Teacher First Name": "TeacherFirstName",
        "Teacher Last Name": "TeacherLastName",
        "Gender": "Gender",
        "Disability": "Disability",
        "Approximate Years Teaching": "ApproximateYearsTeaching",
        "Attendance Rate": "AttendanceRate",
        "80% Attendance": "Percent80Attendance",
        "Statement of Completion": "StatementCompletion",
        "School": "School"
    }
    # Add attended day columns dynamically
    for i in range(1, 16):
        column_mapping[f"Attended Day {i}"] = f"AttendedDay{i}"

    # Root element
    root = ET.Element("ListObject")
    root.set("FirstRow", "2")

    # Set cpdName and cpdYear attributes
    first_data_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    header_to_index = {h: i for i, h in enumerate(headers)}
    root.set("cpdName", str(first_data_row[header_to_index["CPD Name"]]))
    root.set("cpdYear", str(int(first_data_row[header_to_index["Year"]])))

    # Build rows
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if all(cell is None for cell in row):
            continue  # Skip blank rows
        
        row_elem = ET.SubElement(root, "row")
        row_elem.set("Index", str(idx))
        for header, cell_value in zip(headers, row):
            if not header:
                continue  # Skip empty headers
            
            xml_attr = column_mapping.get(header)
            if xml_attr:
                if header in ["Start Date (YYY-MM-DD)", "End Date (YYY-MM-DD)"]:
                    if cell_value is not None:
                        # Convert dates to Excel serial number
                        val = (cell_value - datetime(1899, 12, 30)).days
                    else:
                        val = ""
                else:
                    val = cell_value if cell_value is not None else ""
                
                row_elem.set(xml_attr, str(val).strip())

    return ET.tostring(root, encoding="unicode")



# %%
# Print out one generated XML
somefile = next(iter(all_workbooks.keys()))
print(workbook_to_xml(all_workbooks[somefile]))

# %%
# Code to process the Excel workbook into XML data ready for SQL server stored proc
import uuid
import pyodbc

def extract_cpd_metadata(wb):
    """Extract CPD Name and CPD Year from the workbook."""
    # Access the 'CPD data' sheet
    ws = wb["CPD data"]

    # Read the first data row (row 2)
    first_row = [cell.value for cell in ws[2]]

    # Read headers from row 1
    headers = [cell.value for cell in ws[1]]

    # Create dictionary of {column name -> value}
    row_dict = dict(zip(headers, first_row))

    # Extract CPD Name and Year
    cpd_name = row_dict.get("CPD Name", "")
    cpd_year = row_dict.get("Year", "")

    if not cpd_name or not cpd_year:
        raise ValueError(f"Missing CPD Name or Year in workbook {wb.properties.title}.")

    return cpd_name, cpd_year

import base64

def load_single_workbook_to_sql(file_path, sql_conn):
    wb = all_workbooks[file_path]
    xml_data = workbook_to_xml(wb)
    
    cpd_name, cpd_year = extract_cpd_metadata(wb)
    file_reference = str(uuid.uuid4())
    username = "ghachey@purltek.com"
    cpd_code = cpd_name

    # Encode XML
    xml_base64 = base64.b64encode(xml_data.encode('utf-8')).decode('ascii')

    query = f"""
    DECLARE @p1 XML;
    DECLARE @bin VARBINARY(MAX);

    SET @bin = CAST(CAST('{xml_base64}' AS XML).value('.', 'VARBINARY(MAX)') AS VARBINARY(MAX));
    SET @p1 = CONVERT(XML, @bin);

    EXEC pTeacherWrite.LoadTeacherCpd 
        @cpdData = @p1, 
        @fileReference = '{file_reference}',
        @user = '{username}',
        @cpdCode = '{cpd_code}',
        @cpdYear = {cpd_year};
    """

    cursor = sql_conn.cursor()
    cursor.execute(query)

    # 📌 Important: advance through ALL possible result sets
    while cursor.nextset():
        pass

    sql_conn.commit()
    cursor.close()

    print(f"✅ Successfully loaded {os.path.basename(file_path)} into SQL Server.")


# %%
# Try loading a single workbook in DB
somefile = list(all_workbooks.keys())[1] # change index to load load different one
print(f"Loading file: {os.path.basename(somefile)}")

# Run the loader
load_single_workbook_to_sql(somefile, sql_conn)

# %%
# Bulk load all workbooks to SQL Server
import time

def bulk_load_all_workbooks(all_workbooks, sql_conn, verbose=True):
    total = len(all_workbooks)
    success_count = 0
    fail_count = 0
    start_time = time.time()

    for idx, file_path in enumerate(all_workbooks.keys(), start=1):
        try:
            print(f"({idx}/{total}) Loading: {os.path.basename(file_path)} ...", end=" ")
            load_single_workbook_to_sql(file_path, sql_conn)
            success_count += 1
            if verbose:
                print("✅")
        except Exception as e:
            fail_count += 1
            print(f"❌ Failed: {e}")

    duration = time.time() - start_time
    print("\n=== Bulk Load Summary ===")
    print(f"Total Files Attempted : {total}")
    print(f"✅ Success             : {success_count}")
    print(f"❌ Failed              : {fail_count}")
    print(f"⏱️ Duration           : {duration:.2f} seconds")

# Run it:
bulk_load_all_workbooks(all_workbooks, sql_conn)


# %%
