# ---
# jupyter:
#   jupytext:
#     formats: ipynb,py:percent
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#       jupytext_version: 1.17.0
#   kernelspec:
#     display_name: Python 3 (ipykernel)
#     language: python
#     name: python3
# ---

# %%
# Pull in configuration and set a few things up
import json
import os

import pandas as pd
import random
from datetime import datetime, timedelta
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

def load_config(config_path="config.json"):
    """Load configuration from a JSON file."""
    with open(config_path, 'r') as file:
        config = json.load(file)
    return config['base_url'], config['username'], config['password'], config['output_directory'], config['cpd_directory']

# Test loading configuration
base_url, username, password, output_dir, cpd_directory = load_config()

# Filenames and paths
input_filename = "CPD-source-data-workbook.xlsx"
output_filename = "CPD-source-data-workbook-filled-sample.xlsx"
input_path = os.path.join(cpd_directory, input_filename)
output_path = os.path.join(cpd_directory, output_filename)

print("Configuration loaded successfully.")

# %%
# Extract columns and lists (i.e. used as dropdown validations)

# Use openpyxl to read list values (xlwings doesn't support that easily)
wb_ox = load_workbook(filename=input_path)
list_sheet = wb_ox["Lists"]

# Extract validation lists using header cell position
def extract_list_values_from_header(header_cell):
    col_letter, start_row = coordinate_from_string(header_cell)
    col_index = column_index_from_string(col_letter)
    values = []
    row = start_row + 1
    while list_sheet.cell(row=row, column=col_index).value:
        values.append(list_sheet.cell(row=row, column=col_index).value)
        row += 1
    return values

def extract_list_from_column(col_letter, start_row):
    values = []
    while list_sheet[f"{col_letter}{start_row}"].value:
        values.append(list_sheet[f"{col_letter}{start_row}"].value)
        start_row += 1
    return values

# Load all dropdown lists
cpd_types = extract_list_values_from_header("C4")
cpd_formats = extract_list_values_from_header("E4")
cpd_focuses = extract_list_values_from_header("G4")
genders = extract_list_values_from_header("M4")
schools = extract_list_values_from_header("K4")
years_teaching = extract_list_values_from_header("A15")

print("cpd_types:", cpd_types[:5])
print("cpd_formats:", cpd_formats[:5])
print("cpd_focuses:", cpd_focuses[:5])
print("genders:", genders[:5])
print("schools:", schools[:5])
print("years_teaching:", years_teaching[:5])

# Use xlwings to read and write the CPD Data sheet
wb = xw.Book(input_path)
ws = wb.sheets["CPD data"]
# Read exact column headers from row 1
exact_column_order = ws.range("A1").expand("right").value
print("exact_column_order length:", len(exact_column_order))
print("exact_column_order:", exact_column_order)

# %%
# Get all teachers from EMIS. This depends on notebook teachers.ipynb which needs to run at least once
# to pickle the data locally.
import os
import pickle

cache_dir = "cached-data"

with open(os.path.join(cache_dir, "all_teachers.pkl"), "rb") as f:
    all_teachers = pickle.load(f)

# Optional: filter valid teachers right away
valid_teachers = [
    t for t in all_teachers
    if t['tPayroll'] and t['tSex'] and t['tGiven'] and t['tSurname']
]

print(f"Loaded {len(valid_teachers)} valid teachers from cache.")

# %%
# === Data generation ===
num_teachers = 30
start_date = datetime(2025, 5, 5)  # Monday
duration_days = random.choice([5, 10, 15])
end_date = start_date + timedelta(days=duration_days - 1)
duration_hours = duration_days * 8

fixed_cpd_type = random.choice(cpd_types)
fixed_cpd_format = random.choice(cpd_formats)
fixed_cpd_focus = random.choice(cpd_focuses)
location = "South Tarawa"
year = 2025

def generate_teacher_row(i):    
    use_real_teacher = random.random() < 0.5  # 50% chance

    if use_real_teacher and valid_teachers:
        t = random.choice(valid_teachers)
        pf_number = t['tPayroll']
        first_name = t['tGiven']
        last_name = t['tSurname']
        gender = 'Male' if t['tSex'] == 'M' else 'Female'  # Map to your gender list if needed
    else:
        pf_number = random.randint(1000000, 9999999)
        first_name = f"TeacherFirst{i}"
        last_name = f"TeacherLast{i}"
        gender = random.choice(genders)
        
    disability = "No" if random.random() > 0.1 else "Yes"
    years = random.choice(years_teaching)
    attended_days = ["Yes" if random.random() > 0.1 else "No" for _ in range(duration_days)]    
    attendance_rate = attended_days.count("Yes") / duration_days
    attendance_80 = "Yes" if attendance_rate >= 0.8 else "No"
    completion = "Yes" if attendance_80 == "Yes" else "No"
    school = random.choice(schools)

    base = [
        fixed_cpd_type, fixed_cpd_format, fixed_cpd_focus, location, year,
        start_date.date(), end_date.date(), duration_days, duration_hours,
        pf_number, first_name, last_name, gender, disability, years
    ]
    return base + [attendance_rate, attendance_80, completion, school] + attended_days + [""] * (15 - duration_days)

# Sample row
print("Sample row: ", generate_teacher_row(1))

# %%
# Generate data and load into DataFrame
rows = [generate_teacher_row(i) for i in range(1, num_teachers + 1)]
df = pd.DataFrame(rows, columns=exact_column_order)
display(df[:5])

# %%
# Clear old data and write new data to workbook.
start_cell = "A2"
data_range = ws.range(start_cell).expand("table")
data_range.clear_contents()

# Write the DataFrame values (without header)
ws.range(start_cell).value = df.values.tolist()

# Save and close
wb.save(output_path)
wb.close()

# %%
# %%time
# Generate for many years and worktypes of CPD
# Five years and 8 different types of CPD takes about 3min 27sec on the iMac

# One silent Excel instance for all processing
app = xw.App(visible=False)
app.display_alerts = False
app.screen_updating = False

try:
    for year in range(2020, 2026):  # 2020 to 2025 inclusive
        for cpd_type in cpd_types:
            print(f"Generating for Year: {year}, CPD Type: {cpd_type}")
    
            # === Reopen fresh workbook for each combination ===
            wb = app.books.open(input_path)
            ws = wb.sheets["CPD data"]        
            #wb = xw.Book(input_path)
            #ws = wb.sheets["CPD data"]
            exact_column_order = ws.range("A1").expand("right").value
    
            # === Data generation for this file ===
            num_teachers = 30
            start_date = datetime(year, 5, 5)
            duration_days = random.choice([5, 10, 15])
            end_date = start_date + timedelta(days=duration_days - 1)
            duration_hours = duration_days * 8
    
            fixed_cpd_type = cpd_type
            fixed_cpd_format = random.choice(cpd_formats)
            fixed_cpd_focus = random.choice(cpd_focuses)
            location = "South Tarawa"
            year = year
    
            def generate_teacher_row(i):
                use_real_teacher = random.random() < 0.5  # 50% chance
    
                if use_real_teacher and valid_teachers:
                    t = random.choice(valid_teachers)
                    pf_number = t['tPayroll']
                    first_name = t['tGiven']
                    last_name = t['tSurname']
                    gender = 'Male' if t['tSex'] == 'M' else 'Female'  # Map to your gender list if needed
                else:
                    pf_number = random.randint(1000000, 9999999)
                    first_name = f"TeacherFirst{i}"
                    last_name = f"TeacherLast{i}"
                    gender = random.choice(genders)
            
                disability = "No" if random.random() > 0.1 else "Yes"
                years_taught = random.choice(years_teaching)
                attended_days = ["Yes" if random.random() > 0.1 else "No" for _ in range(duration_days)]
                attendance_rate = attended_days.count("Yes") / duration_days
                attendance_80 = "Yes" if attendance_rate >= 0.8 else "No"
                completion = "Yes" if attendance_80 == "Yes" else "No"
                school = random.choice(schools)
    
                base = [
                    fixed_cpd_type, fixed_cpd_format, fixed_cpd_focus, location, year,
                    start_date.date(), end_date.date(), duration_days, duration_hours,
                    pf_number, first_name, last_name, gender, disability, years_taught
                ]
                return base + [attendance_rate, attendance_80, completion, school] + attended_days + [""] * (15 - duration_days)
    
            rows = [generate_teacher_row(i) for i in range(1, num_teachers + 1)]
            df = pd.DataFrame(rows, columns=exact_column_order)
    
            # Write to sheet
            start_cell = "A2"
            ws.range(start_cell).expand("table").clear_contents()
            ws.range(start_cell).value = df.values.tolist()
    
            # Save under new name
            safe_cpd_name = fixed_cpd_type.replace(" ", "_").replace("/", "_")
            filename = f"CPD-{safe_cpd_name}-{year}.xlsx"
            save_path = os.path.join(cpd_directory, filename)
            wb.save(save_path)
            wb.close()
finally:
    app.quit()  # Ensures Excel closes even on error

# %%
